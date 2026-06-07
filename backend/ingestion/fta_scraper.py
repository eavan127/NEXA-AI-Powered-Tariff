import asyncio
import re
import time
import httpx
from datetime import datetime, timezone
from supabase import create_client

from config import settings
from ingestion.change_detector import has_source_changed, store_source_hash
from ingestion.pdf_extractor import extract_fta_rates_from_pdf

# ── JKDM ─────────────────────────────────────────────────────────
# Only our 5 target HS codes (4-digit for JKDM search)
# NOTE: 8501.52 replaced by 8501.10 (spindle motors, 8501103000)
# NOTE: 8542.31 (Taiwan) not scraped from JKDM — Taiwan has no FTA with Malaysia
HS_CODES_TO_SCRAPE = ["8534", "7604", "8501", "9001"]

# Maps JKDM dropdown label → fta_name stored in DB
# ATIGA is NOT in JKDM dropdown — handled separately via PDF fallback
# Only scrape FTAs relevant to our 4 origin countries
JKDM_TARIFF_MAP = {
    "PDK 2025 - PERINTAH DUTI KASTAM 2025":              "PDK",
    "ACFTA - ASEAN CHINA FREE TRADE AGREEMENT":           "ACFTA",
    "AKFTA - ASEAN KOREA FREE TRADE AGREEMENT":           "AKFTA",
    "RCEP - REGIONAL COMPREHENSIVE ECONOMIC PARTNERSHIP": "RCEP",
}

# For each FTA+HS searched, which 10-digit sub-code row to read
# JKDM returns multiple sub-codes per header; we pick the specific one
# that matches our product type
JKDM_SUBCODE_TARGET = {
    "8534": "8534001000",   # single-sided PCB
    "7604": "7604291000",   # extruded bars and rods (aluminium alloy profile)
    "8501": "8501103000",   # spindle motors (DC electric)
    "9001": "9001901000",   # optical lens for cameras/projectors
    "8542": "8542310000",   # electronic ICs / processors (Taiwan MFN baseline)
}

# Origin country to assign per FTA when upserting (JKDM has no origin selector)
JKDM_FTA_ORIGIN = {
    "PDK":   "MFN",           # MFN baseline — applies to all countries
    "ACFTA": "China",
    "AKFTA": "South Korea",
    "RCEP":  None,            # multiple origins — see RCEP_COUNTRIES below
}
RCEP_COUNTRIES = ["Vietnam", "China", "South Korea"]

# Only search relevant HS codes per FTA — avoids wasted requests + wrong insertions
# ACFTA is China-only → only motors (8501); AKFTA is Korea-only → only optical (9001)
# RCEP and PDK cover all 4 HS codes
JKDM_FTA_HS_SCOPE = {
    "PDK":   ["8534", "7604", "8501", "9001", "8542"],  # MFN baseline — all 5 HS codes
    "RCEP":  ["8534", "7604", "8501", "9001"],           # applies to Vietnam/China/S.Korea
    "ACFTA": ["8501"],                                   # China motors only
    "AKFTA": ["9001"],                                   # Korea optical only
}

# ── MITI page URLs — restricted to our 4 FTAs only ───────────────
# ATIGA included here as PDF fallback source (JKDM doesn't have ATIGA)
# Other FTAs excluded — not needed for our 4 origin countries demo
FTA_PAGES = {
    "ATIGA": "https://fta.miti.gov.my/index.php/pages/view/asean-afta",
    "ACFTA": "https://fta.miti.gov.my/index.php/pages/view/asean-china",
    "AKFTA": "https://fta.miti.gov.my/index.php/pages/view/asean-korea",
    "RCEP":  "https://fta.miti.gov.my/index.php/pages/view/rcep",
}

# ── Complete metadata for all 17 FTAs ────────────────────────────
FTA_COVERAGE_META = {
    "AKFTA": {
        "effective_date": "2007-06-01", "in_force_date": "2007-06-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/asean-korea",
        "description":    "ASEAN-Korea Free Trade Agreement",
        "roo_summary":    "RVC >= 40% (build-up) OR CTH",
    },
    "ATIGA": {
        "effective_date": "2010-01-01", "in_force_date": "2010-01-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/asean-afta",
        "description":    "ASEAN Trade in Goods Agreement (ATIGA)",
        "roo_summary":    "RVC >= 40% (ATIGA build-up) OR CTH at 6-digit level",
    },
    "AHKFTA": {
        "effective_date": "2019-06-17", "in_force_date": "2019-06-17",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/asean-hongkong-china",
        "description":    "ASEAN-Hong Kong Free Trade Agreement",
        "roo_summary":    "RVC >= 35% OR CTH",
    },
    "ACFTA": {
        "effective_date": "2005-07-20", "in_force_date": "2005-07-20",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/asean-china",
        "description":    "ASEAN-China Free Trade Agreement",
        "roo_summary":    "RVC >= 40% OR CTH at subheading level",
    },
    "MICECA": {
        "effective_date": "2012-07-01", "in_force_date": "2012-07-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/malaysia-india",
        "description":    "Malaysia-India Comprehensive Economic Cooperation Agreement",
        "roo_summary":    "RVC >= 35% OR CTH",
    },
    "AIFTA": {
        "effective_date": "2010-01-01", "in_force_date": "2010-01-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/asean-india",
        "description":    "ASEAN-India Free Trade Agreement",
        "roo_summary":    "RVC >= 35% OR CTH; India has Sensitive Track exclusions",
    },
    "MPCEPA": {
        "effective_date": "2008-01-01", "in_force_date": "2008-01-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/malaysia-pakistan",
        "description":    "Malaysia-Pakistan Closer Economic Partnership Agreement",
        "roo_summary":    "RVC >= 40%",
    },
    "MCFTA": {
        "effective_date": "2012-02-25", "in_force_date": "2012-02-25",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/malaysia-chile",
        "description":    "Malaysia-Chile Free Trade Agreement",
        "roo_summary":    "RVC >= 40% OR CTH",
    },
    "AJCEP": {
        "effective_date": "2008-12-01", "in_force_date": "2008-12-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/asean-japan",
        "description":    "ASEAN-Japan Comprehensive Economic Partnership",
        "roo_summary":    "RVC >= 40% OR CTH; compare with MJEPA, use lower rate",
    },
    "MNZFTA": {
        "effective_date": "2010-08-01", "in_force_date": "2010-08-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/malaysia-newzealand",
        "description":    "Malaysia-New Zealand Free Trade Agreement",
        "roo_summary":    "RVC >= 40%; NZ also covered by AANZFTA and CPTPP",
    },
    "RCEP": {
        "effective_date": "2022-01-01", "in_force_date": "2022-01-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/rcep",
        "description":    "Regional Comprehensive Economic Partnership",
        "roo_summary":    "RVC >= 40% (build-up) OR CTSH; cumulation allowed across all RCEP members",
    },
    "TPS-OIC": {
        "effective_date": "1991-01-01", "in_force_date": "1991-01-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/109",
        "description":    "Trade Preferential System — Organisation of Islamic Cooperation",
        "roo_summary":    "RVC >= 40%; margin of preference, not zero rate",
    },
    "MTFTA": {
        "effective_date": "2015-08-01", "in_force_date": "2015-08-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/malaysia-turkey",
        "description":    "Malaysia-Turkey Free Trade Agreement",
        "roo_summary":    "RVC >= 40% OR CTH; some goods staging through 2026-2028",
    },
    "MJEPA": {
        "effective_date": "2008-07-13", "in_force_date": "2008-07-13",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/malaysia-japan",
        "description":    "Malaysia-Japan Economic Partnership Agreement",
        "roo_summary":    "More generous than AJCEP for some goods; always check both",
    },
    "AANZFTA": {
        "effective_date": "2012-01-01", "in_force_date": "2012-01-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/asean-australia-newzealand",
        "description":    "ASEAN-Australia-New Zealand Free Trade Agreement",
        "roo_summary":    "RVC >= 40% OR CTH; AU also covered by MAFTA and RCEP",
    },
    "CPTPP": {
        "effective_date": "2022-11-29", "in_force_date": "2022-11-29",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/tpp_cptpp",
        "description":    "Comprehensive and Progressive Agreement for Trans-Pacific Partnership",
        "roo_summary":    "RVC >= 40% OR CTH; categories EIF/B5/B7/B10/B12/X",
    },
    "MAFTA": {
        "effective_date": "2013-01-01", "in_force_date": "2013-01-01",
        "source_url":     "https://fta.miti.gov.my/index.php/pages/view/malaysia-australia",
        "description":    "Malaysia-Australia Free Trade Agreement",
        "roo_summary":    "RVC >= 40%; most goods at 0%; AU also covered by AANZFTA, RCEP, CPTPP",
    },
}

FTA_MEMBERS = {
    "ATIGA":      ["Malaysia", "Indonesia", "Thailand", "Singapore",
                   "Philippines", "Vietnam", "Brunei", "Cambodia", "Laos", "Myanmar"],
    "ACFTA":      ["China", "Malaysia", "Indonesia", "Thailand",
                   "Singapore", "Philippines", "Vietnam", "Brunei", "Cambodia", "Laos", "Myanmar"],
    "AKFTA":      ["South Korea", "Malaysia", "Indonesia", "Thailand",
                   "Singapore", "Philippines", "Vietnam", "Brunei", "Cambodia", "Laos", "Myanmar"],
    "AJCEP":      ["Japan", "Malaysia", "Indonesia", "Thailand",
                   "Singapore", "Philippines", "Vietnam", "Brunei", "Cambodia", "Laos", "Myanmar"],
    "AIFTA":      ["India", "Malaysia", "Indonesia", "Thailand",
                   "Singapore", "Philippines", "Vietnam", "Brunei", "Cambodia", "Laos", "Myanmar"],
    "AANZFTA":    ["Australia", "New Zealand", "Malaysia", "Indonesia", "Thailand",
                   "Singapore", "Philippines", "Vietnam", "Brunei", "Cambodia", "Laos", "Myanmar"],
    "AHKFTA":     ["Hong Kong", "Malaysia", "Indonesia", "Thailand",
                   "Singapore", "Philippines", "Vietnam", "Brunei", "Cambodia", "Laos", "Myanmar"],
    "RCEP":       ["China", "Japan", "South Korea", "Australia", "New Zealand",
                   "Malaysia", "Indonesia", "Thailand", "Singapore", "Philippines",
                   "Vietnam", "Brunei", "Cambodia", "Laos", "Myanmar"],
    "CPTPP":      ["Australia", "Brunei", "Canada", "Chile", "Japan", "Malaysia",
                   "Mexico", "New Zealand", "Peru", "Singapore", "Vietnam", "United Kingdom"],
    "MAFTA":      ["Malaysia", "Australia"],
    "MCFTA":      ["Malaysia", "Chile"],
    "MICECA":     ["Malaysia", "India"],
    "MJEPA":      ["Malaysia", "Japan"],
    "MNZFTA":     ["Malaysia", "New Zealand"],
    "MPCEPA":     ["Malaysia", "Pakistan"],
    "MTFTA":      ["Malaysia", "Turkey"],
    "TPS-OIC":    ["Malaysia", "Turkey", "Pakistan", "Bangladesh", "Indonesia",
                   "Iran", "Egypt", "Morocco", "Nigeria", "Saudi Arabia",
                   "Jordan", "Tunisia", "Senegal", "Cameroon", "Burkina Faso"],
}

COUNTRY_CODE_MAP = {
    "CHN": "China",       "JPN": "Japan",         "KOR": "South Korea",
    "AUS": "Australia",   "NZL": "New Zealand",   "IND": "India",
    "CHL": "Chile",       "PAK": "Pakistan",       "TUR": "Turkey",
    "IDN": "Indonesia",   "THA": "Thailand",       "SGP": "Singapore",
    "PHL": "Philippines", "VNM": "Vietnam",        "BRN": "Brunei",
    "KHM": "Cambodia",    "LAO": "Laos",           "MMR": "Myanmar",
    "HKG": "Hong Kong",   "CAN": "Canada",         "MEX": "Mexico",
    "GBR": "United Kingdom", "PER": "Peru",        "BGD": "Bangladesh",
}

# Only RCEP has a known directory listing URL for PDFs
# ATIGA PDF used as fallback when JKDM is unavailable (MITI site may be down)
MITI_PDF_DIRS = {
    "RCEP": "https://fta.miti.gov.my/miti-fta/resources/RCEP/",
}


# ════════════════════════════════════════════════════════════════
# JKDM — sync Playwright wrapped in run_in_executor
# ════════════════════════════════════════════════════════════════

async def scrape_jkdm_rates() -> dict:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _scrape_jkdm_sync)


def _scrape_jkdm_sync() -> dict:
    """
    Scrape JKDM ezhs.customs.gov.my for real FTA rates.
    - Searches by 4-digit HS header (JKDM only accepts 4 digits)
    - Finds the specific 10-digit sub-code row we care about (JKDM_SUBCODE_TARGET)
    - Reads CURRENT RATE from that row
    - Stores with correct origin_country per FTA
    - ATIGA not in JKDM — handled by _scrape_atiga_fallback()
    """
    from playwright.sync_api import sync_playwright
    supabase = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
    rates_updated = 0

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True, args=["--ignore-certificate-errors"]
            )
            context = browser.new_context(ignore_https_errors=True)
            page    = context.new_page()

            for tariff_label, fta_name in JKDM_TARIFF_MAP.items():
                hs_scope = JKDM_FTA_HS_SCOPE.get(fta_name, [])
                for hs_4digit in hs_scope:
                    target_subcode = JKDM_SUBCODE_TARGET.get(hs_4digit)
                    if not target_subcode:
                        continue

                    for attempt in range(3):
                        try:
                            page.goto("https://ezhs.customs.gov.my", timeout=60000)
                            page.wait_for_load_state("domcontentloaded")
                            page.select_option("select", label=tariff_label)
                            time.sleep(0.5)
                            page.fill("input[type='text']", hs_4digit)
                            page.click("input[type='submit'], button[type='submit']")
                            # Wait for results — gracefully handle pages with no table
                            try:
                                page.wait_for_selector("table tr td", timeout=15000)
                            except Exception:
                                # No table appeared (no-results page or slow load) — wait fixed delay
                                time.sleep(3)
                            time.sleep(1)

                            rate = _extract_rate_for_subcode(page, target_subcode)

                            if rate is not None:
                                hs_dot  = _subcode_to_dot(target_subcode)
                                origins = RCEP_COUNTRIES if fta_name == "RCEP" else [JKDM_FTA_ORIGIN.get(fta_name, "Unknown")]

                                for origin in origins:
                                    supabase.table("fta_rates").upsert(
                                        {
                                            "fta_name":              fta_name,
                                            "hs_code":               hs_dot,
                                            "preferential_rate_pct": rate,
                                            "origin_country":        origin,
                                            "effective_date":        "2025-01-01",
                                            "pdf_source_url":        "https://ezhs.customs.gov.my",
                                            "last_verified_at":      datetime.now(timezone.utc).isoformat(),
                                        },
                                        on_conflict="fta_name,hs_code,origin_country"
                                    ).execute()
                                    rates_updated += 1

                                print(f"[jkdm] {fta_name} | {hs_dot} ({target_subcode}) → {rate}% [{', '.join(origins)}]")

                                supabase.table("pipeline_logs").insert({
                                    "event_type":  "jkdm_rate_scraped",
                                    "source_url":  "https://ezhs.customs.gov.my",
                                    "description": f"{fta_name} | HS {hs_dot} | sub-code {target_subcode} = {rate}%",
                                    "detail":      {"fta": fta_name, "hs_code": hs_dot,
                                                    "subcode": target_subcode, "rate": rate},
                                }).execute()
                            else:
                                print(f"[jkdm] Sub-code {target_subcode} not found in {fta_name} results")
                            break

                        except Exception as e:
                            print(f"[jkdm] {fta_name}/{hs_4digit} attempt {attempt+1}: {e}")
                            time.sleep(2 ** attempt)

            browser.close()

    except Exception as e:
        print(f"[jkdm] _scrape_jkdm_sync failed: {e}")

    # After JKDM: ATIGA Vietnam rates from MITI Excel
    try:
        _scrape_atiga_fallback(supabase)
    except Exception as e:
        print(f"[atiga] scrape failed (non-fatal): {e}")

    # ACFTA China — Malaysia tariff schedule + RoO PSR PDFs
    try:
        _scrape_acfta_malaysia(supabase)
    except Exception as e:
        print(f"[acfta] scrape failed (non-fatal): {e}")

    # AKFTA Korea PSR PDF — cross-check RoO rules
    try:
        _scrape_akfta_psr(supabase)
    except Exception as e:
        print(f"[akfta] PSR scrape failed (non-fatal): {e}")

    return {"rates_updated": rates_updated}


def _extract_rate_for_subcode(page, target_subcode: str):
    """
    Parse the JKDM results table.

    PDK format:  HEADER(0) | SUB(1) | ITEM(2) | UNIT(3) | DESCRIPTION(4) | IMPORT RATE(5)
    FTA format:  ... | FULL_HS(3) | ... | CURRENT RATE(5)

    Detection: PDK rows have a 4-digit number in col0, so we reconstruct
    full_hs = col0 + col1.zfill(2) + col2.zfill(4).
    """
    try:
        rows = page.query_selector_all("table tr")

        for row in rows:
            cells = row.query_selector_all("td")
            if len(cells) < 6:
                continue

            c0 = cells[0].inner_text().strip().replace(" ", "")
            c1 = cells[1].inner_text().strip().replace(" ", "")
            c2 = cells[2].inner_text().strip().replace(" ", "")

            # PDK format: col0 is 4-digit header, col1 is 2-digit sub, col2 is 4-digit item
            if c0.isdigit() and len(c0) == 4 and c1.isdigit() and c2.isdigit():
                full_hs   = c0 + c1.zfill(2) + c2.zfill(4)
                rate_text = cells[5].inner_text().strip().replace("%", "").replace(",", ".").strip()
            else:
                # FTA format: full HS code in col3
                full_hs   = cells[3].inner_text().strip().replace(" ", "")
                rate_text = cells[5].inner_text().strip().replace("%", "").replace(",", ".").strip()

            if rate_text in ("", "-", "N/A", "CURRENT RATE", "IMPORT RATE"):
                continue

            if full_hs != target_subcode:
                continue

            if rate_text.lower() in ("free", "nil", "0"):
                return 0.0
            try:
                rate = float(rate_text)
                if 0.0 <= rate <= 100.0:
                    return rate
            except ValueError:
                pass

        return None
    except Exception as e:
        print(f"[jkdm] _extract_rate_for_subcode({target_subcode}) error: {e}")
        return None


def _subcode_to_dot(subcode: str) -> str:
    """Convert 10-digit JKDM code to 6-digit dot notation: 8534001000 → 8534.00"""
    digits = subcode[:6]  # take first 6 digits
    return f"{digits[:4]}.{digits[4:6]}"


def _scrape_atiga_fallback(supabase) -> None:
    """
    ATIGA is NOT in JKDM dropdown.
    Primary: visit MITI ATIGA page, find Vietnam Excel tariff schedule, download + parse.
    Fallback: use known ASEAN full-liberalisation rates if download fails.
    """
    import io
    import re
    import tempfile
    import requests
    import openpyxl

    ATIGA_PAGE   = "https://fta.miti.gov.my/index.php/pages/view/asean-afta"
    ORIGIN       = "Vietnam"
    # 4-digit HS prefixes we care about (dot-notation lookup keys)
    TARGET_HS = {
        "8534": "8534.00",
        "7604": "7604.29",
        "8501": "8501.10",
        "9001": "9001.90",
    }
    FALLBACK_RATES = {
        "8534.00": 0.0,   # electronics — ASEAN fully liberalised since 2010
        "7604.29": 0.0,   # aluminium  — ASEAN fully liberalised since 2015
        "8501.10": 0.0,   # motors     — ASEAN fully liberalised
        "9001.90": 0.0,   # optics     — ASEAN fully liberalised
    }

    excel_url  = None
    excel_file = None

    # ── Step 1: find the Vietnam Excel link on MITI ATIGA page ──────
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--ignore-certificate-errors"])
            page    = browser.new_context(ignore_https_errors=True).new_page()
            page.goto(ATIGA_PAGE, timeout=30000)
            try:
                page.wait_for_selector("a", timeout=10000)
            except Exception:
                pass

            # Find link whose text contains "Viet Nam" or "Vietnam" + "Annex 2"
            links = page.query_selector_all("a")
            for link in links:
                text = (link.inner_text() or "").strip()
                href = link.get_attribute("href") or ""
                if re.search(r"viet.?nam", text, re.IGNORECASE) and re.search(r"annex.?2|tariff.?schedule", text, re.IGNORECASE):
                    excel_url = href if href.startswith("http") else f"https://fta.miti.gov.my{href}"
                    print(f"[atiga] Found Vietnam Excel link: {excel_url}")
                    break

            browser.close()
    except Exception as e:
        print(f"[atiga] Playwright failed finding Excel link: {e}")

    # ── Step 2: download and parse the Excel ────────────────────────
    if excel_url:
        try:
            resp = requests.get(excel_url, timeout=60, verify=False)
            resp.raise_for_status()
            excel_file = io.BytesIO(resp.content)
            print(f"[atiga] Downloaded Excel ({len(resp.content)//1024} KB)")
        except Exception as e:
            print(f"[atiga] Excel download failed: {e}")

    scraped_rates = {}
    if excel_file:
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    # Find column with HS code (first cell that looks like a number/string HS)
                    for i, cell in enumerate(row):
                        if cell is None:
                            continue
                        cell_str = str(cell).strip().replace(".", "").replace(" ", "")
                        if not cell_str.isdigit():
                            continue
                        prefix = cell_str[:4]
                        if prefix not in TARGET_HS:
                            continue
                        # Rate is typically in a nearby numeric column — scan rightward
                        for j in range(i + 1, min(i + 8, len(row))):
                            rate_val = row[j]
                            if rate_val is None:
                                continue
                            try:
                                rate = float(str(rate_val).replace("%", "").strip())
                                if 0.0 <= rate <= 100.0:
                                    hs_dot = TARGET_HS[prefix]
                                    if hs_dot not in scraped_rates:
                                        scraped_rates[hs_dot] = rate
                                    break
                            except (ValueError, TypeError):
                                pass
            wb.close()
            print(f"[atiga] Parsed rates from Excel: {scraped_rates}")
        except Exception as e:
            print(f"[atiga] Excel parse failed: {e}")

    # ── Step 3: upsert — use scraped if available, else fallback ────
    source_url  = excel_url or "https://fta.miti.gov.my (Excel unavailable — fallback)"
    needs_review = not bool(scraped_rates)

    for hs_dot, fallback_rate in FALLBACK_RATES.items():
        rate = scraped_rates.get(hs_dot, fallback_rate)
        try:
            supabase.table("fta_rates").upsert(
                {
                    "fta_name":              "ATIGA",
                    "hs_code":               hs_dot,
                    "preferential_rate_pct": rate,
                    "origin_country":        ORIGIN,
                    "effective_date":        "2010-01-01",
                    "pdf_source_url":        source_url,
                    "needs_review":          needs_review,
                    "last_verified_at":      datetime.now(timezone.utc).isoformat(),
                },
                on_conflict="fta_name,hs_code,origin_country"
            ).execute()
            src = "Excel" if hs_dot in scraped_rates else "fallback"
            print(f"[atiga] ATIGA {hs_dot} Vietnam = {rate}% [{src}]")
        except Exception as e:
            print(f"[atiga] Upsert failed for {hs_dot}: {e}")

    supabase.table("pipeline_logs").insert({
        "event_type":  "atiga_excel_scraped" if scraped_rates else "atiga_fallback_used",
        "source_url":  source_url,
        "description": f"ATIGA Vietnam rates: {len(scraped_rates)} from Excel, {len(FALLBACK_RATES)-len(scraped_rates)} fallback",
        "detail":      {"scraped": scraped_rates, "excel_url": excel_url},
    }).execute()


def _scrape_acfta_malaysia(supabase) -> None:
    """
    Visit MITI ACFTA page → Related Documents tab → Malaysia tariff schedule.
    Downloads the Malaysia file (Excel/PDF), extracts rate for HS 8501.10 (motors from China).
    Also grabs the two Revised ROO PDFs for ACFTA PSR cross-check.
    General RoO confirmed from page: RVC >= 40% (build-down) or PSR.
    """
    import io
    import re
    import requests
    import openpyxl
    import pdfplumber

    ACFTA_PAGE = "https://fta.miti.gov.my/index.php/pages/view/asean-china"
    TARGET_HS  = {"8501": "8501.10"}   # only China motors for our project

    malaysia_url = None
    roo_pdf_urls = []

    # ── Step 1: find Malaysia tariff schedule link + RoO PDFs ─────
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--ignore-certificate-errors"])
            page    = browser.new_context(ignore_https_errors=True).new_page()
            page.goto(ACFTA_PAGE, timeout=30000)

            # Click "Related Documents" tab if present
            try:
                page.click("text=Related Documents", timeout=5000)
                page.wait_for_selector("a", timeout=5000)
            except Exception:
                pass

            for link in page.query_selector_all("a"):
                text = (link.inner_text() or "").strip()
                href = link.get_attribute("href") or ""
                full = href if href.startswith("http") else f"https://fta.miti.gov.my{href}"

                # Malaysia tariff schedule
                if re.search(r"\bmalaysia\b", text, re.IGNORECASE) and not malaysia_url:
                    if re.search(r"\.(xls|xlsx|pdf|zip)$", href, re.IGNORECASE) or re.search(r"tariff|schedule|annex", text, re.IGNORECASE):
                        malaysia_url = full
                        print(f"[acfta] Found Malaysia schedule: {malaysia_url}")

                # Revised ROO PDFs (both links shown on the page)
                if re.search(r"ROO|Rules.of.Origin|Annex", text, re.IGNORECASE) and href.endswith(".pdf"):
                    roo_pdf_urls.append(full)

            browser.close()
    except Exception as e:
        print(f"[acfta] Playwright failed: {e}")

    # ── Step 2: download + parse Malaysia tariff schedule ─────────
    scraped_rates = {}
    if malaysia_url:
        try:
            resp = requests.get(malaysia_url, timeout=60, verify=False)
            resp.raise_for_status()
            content = resp.content
            print(f"[acfta] Downloaded Malaysia schedule ({len(content)//1024} KB)")

            # Try Excel first
            if malaysia_url.lower().endswith((".xls", ".xlsx")) or b"xl/" in content[:2000]:
                try:
                    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                    for ws in wb.worksheets:
                        for row in ws.iter_rows(values_only=True):
                            for i, cell in enumerate(row):
                                cell_str = str(cell or "").strip().replace(".", "").replace(" ", "")
                                if not cell_str.isdigit():
                                    continue
                                prefix = cell_str[:4]
                                if prefix not in TARGET_HS:
                                    continue
                                for j in range(i + 1, min(i + 8, len(row))):
                                    try:
                                        rate = float(str(row[j] or "").replace("%", "").strip())
                                        if 0.0 <= rate <= 100.0:
                                            scraped_rates[TARGET_HS[prefix]] = rate
                                            break
                                    except (ValueError, TypeError):
                                        pass
                    wb.close()
                except Exception as e:
                    print(f"[acfta] Excel parse failed: {e}")

            # Fallback: PDF parse
            if not scraped_rates:
                try:
                    with pdfplumber.open(io.BytesIO(content)) as pdf:
                        for pg in pdf.pages:
                            for table in (pg.extract_tables() or []):
                                for row in table:
                                    row_text = " ".join(str(c) for c in row if c)
                                    for prefix, hs_dot in TARGET_HS.items():
                                        if prefix in row_text and hs_dot not in scraped_rates:
                                            m = re.search(r"(\d+(?:\.\d+)?)\s*%", row_text)
                                            if m:
                                                scraped_rates[hs_dot] = float(m.group(1))
                except Exception as e:
                    print(f"[acfta] PDF parse failed: {e}")

            print(f"[acfta] Parsed rates: {scraped_rates}")
        except Exception as e:
            print(f"[acfta] Schedule download failed: {e}")

    # ── Step 3: upsert scraped ACFTA rates as cross-check ─────────
    for hs_dot, rate in scraped_rates.items():
        try:
            existing = (supabase.table("fta_rates")
                .select("preferential_rate_pct")
                .eq("fta_name", "ACFTA").eq("hs_code", hs_dot).eq("origin_country", "China")
                .limit(1).execute())
            if existing.data:
                old = existing.data[0]["preferential_rate_pct"]
                if old != rate:
                    print(f"[acfta] RATE CHANGED: ACFTA {hs_dot} China {old}% → {rate}%")
                    supabase.table("pipeline_logs").insert({
                        "event_type":  "acfta_rate_changed",
                        "source_url":  malaysia_url,
                        "description": f"ACFTA {hs_dot} China changed: {old}% → {rate}%",
                        "detail":      {"hs_code": hs_dot, "old": old, "new": rate},
                    }).execute()

            supabase.table("fta_rates").upsert({
                "fta_name":              "ACFTA",
                "hs_code":               hs_dot,
                "preferential_rate_pct": rate,
                "origin_country":        "China",
                "effective_date":        "2005-07-20",
                "pdf_source_url":        malaysia_url,
                "needs_review":          False,
                "last_verified_at":      datetime.now(timezone.utc).isoformat(),
            }, on_conflict="fta_name,hs_code,origin_country").execute()
            print(f"[acfta] ACFTA {hs_dot} China = {rate}% [from Malaysia schedule]")
        except Exception as e:
            print(f"[acfta] Upsert failed {hs_dot}: {e}")

    # ── Step 4: parse RoO PDFs → upsert into roo_rules ────────────
    for pdf_url in roo_pdf_urls[:2]:   # max 2 PDFs
        try:
            resp = requests.get(pdf_url, timeout=60, verify=False)
            resp.raise_for_status()
            with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
                for pg in pdf.pages:
                    for table in (pg.extract_tables() or []):
                        for row in table:
                            row_text = " ".join(str(c) for c in row if c)
                            for prefix, hs_dot in TARGET_HS.items():
                                if prefix not in row_text:
                                    continue
                                roo_type = "RVC_OR_CTH"
                                rvc_pct  = 40.0
                                if re.search(r"CC\b", row_text):
                                    roo_type = "CC"
                                elif re.search(r"CTH", row_text):
                                    roo_type = "RVC_OR_CTH"
                                m = re.search(r"(\d+)\s*%", row_text)
                                if m:
                                    rvc_pct = float(m.group(1))
                                supabase.table("roo_rules").upsert({
                                    "fta_name":          "ACFTA",
                                    "hs_code":           hs_dot,
                                    "roo_type":          roo_type,
                                    "rvc_threshold_pct": rvc_pct,
                                    "rule_text":         row_text[:300],
                                    "source_url":        pdf_url,
                                }, on_conflict="fta_name,hs_code").execute()
                                print(f"[acfta] RoO {hs_dot}: {roo_type} {rvc_pct}% (PSR PDF)")
        except Exception as e:
            print(f"[acfta] RoO PDF {pdf_url} failed: {e}")

    supabase.table("pipeline_logs").insert({
        "event_type":  "acfta_malaysia_scraped",
        "source_url":  malaysia_url or ACFTA_PAGE,
        "description": f"ACFTA Malaysia schedule: {len(scraped_rates)} rates, {len(roo_pdf_urls)} RoO PDFs processed",
        "detail":      {"rates": scraped_rates, "roo_pdfs": roo_pdf_urls},
    }).execute()


def _scrape_akfta_psr(supabase) -> None:
    """
    Visit MITI AKFTA page, find "AKFTA PSRs in HS 2012" PDF link, download it,
    extract Product-Specific Rules for our target HS codes, upsert into roo_rules.
    Used as cross-check against hardcoded RoO values.
    RoO general rule confirmed from page: RVC >= 40% (build-down or build-up) OR CTH.
    """
    import io
    import re
    import pdfplumber
    import requests

    AKFTA_PAGE  = "https://fta.miti.gov.my/index.php/pages/view/asean-korea"
    TARGET_HS   = {"8534", "7604", "8501", "9001"}   # 4-digit prefixes we care about
    HS_DOT_MAP  = {
        "8534": "8534.00", "7604": "7604.29",
        "8501": "8501.10", "9001": "9001.90",
    }

    pdf_url  = None
    pdf_bytes = None

    # ── Step 1: find PSR PDF link ─────────────────────────────────
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--ignore-certificate-errors"])
            page    = browser.new_context(ignore_https_errors=True).new_page()
            page.goto(AKFTA_PAGE, timeout=30000)
            try:
                page.wait_for_selector("a", timeout=10000)
            except Exception:
                pass

            for link in page.query_selector_all("a"):
                text = (link.inner_text() or "").strip()
                href = link.get_attribute("href") or ""
                if re.search(r"PSR|product.specific.rule", text, re.IGNORECASE):
                    pdf_url = href if href.startswith("http") else f"https://fta.miti.gov.my{href}"
                    print(f"[akfta] Found PSR PDF link: {pdf_url}")
                    break
            browser.close()
    except Exception as e:
        print(f"[akfta] Playwright failed finding PSR link: {e}")

    # ── Step 2: download PDF ──────────────────────────────────────
    if pdf_url:
        try:
            resp = requests.get(pdf_url, timeout=60, verify=False)
            resp.raise_for_status()
            pdf_bytes = resp.content
            print(f"[akfta] Downloaded PSR PDF ({len(pdf_bytes)//1024} KB)")
        except Exception as e:
            print(f"[akfta] PDF download failed: {e}")

    if not pdf_bytes:
        print("[akfta] PSR PDF unavailable — skipping RoO cross-check")
        return

    # ── Step 3: parse PDF for PSR rows ───────────────────────────
    # PSR table rows look like: "8534.00  |  RVC 40% or CTH"
    # We look for our HS prefixes then read the rule text from nearby cells.
    psr_found = {}
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for pg in pdf.pages:
                tables = pg.extract_tables()
                for table in tables:
                    for row in table:
                        if not row:
                            continue
                        row_text = " ".join(str(c) for c in row if c)
                        for prefix in TARGET_HS:
                            if prefix in row_text and prefix not in psr_found:
                                # Determine rule type from text
                                roo_type  = "RVC_OR_CTH"
                                rvc_pct   = 40.0
                                rule_text = row_text.strip()
                                if re.search(r"wholly\s+obtain", rule_text, re.IGNORECASE):
                                    roo_type = "WO"
                                elif re.search(r"CTH", rule_text):
                                    roo_type = "RVC_OR_CTH"
                                elif re.search(r"CC\b", rule_text):
                                    roo_type = "CC"
                                m = re.search(r"(\d+)\s*%", rule_text)
                                if m:
                                    rvc_pct = float(m.group(1))
                                psr_found[prefix] = {
                                    "roo_type":         roo_type,
                                    "rvc_threshold_pct": rvc_pct,
                                    "rule_text":         rule_text[:300],
                                }
    except Exception as e:
        print(f"[akfta] PDF parse failed: {e}")

    # ── Step 4: upsert into roo_rules ────────────────────────────
    for prefix, rule in psr_found.items():
        hs_dot = HS_DOT_MAP.get(prefix, f"{prefix[:4]}.{prefix[4:6]}")
        try:
            supabase.table("roo_rules").upsert(
                {
                    "fta_name":           "AKFTA",
                    "hs_code":            hs_dot,
                    "roo_type":           rule["roo_type"],
                    "rvc_threshold_pct":  rule["rvc_threshold_pct"],
                    "rule_text":          rule["rule_text"],
                    "source_url":         pdf_url,
                },
                on_conflict="fta_name,hs_code",
            ).execute()
            print(f"[akfta] RoO {hs_dot}: {rule['roo_type']} {rule['rvc_threshold_pct']}% (from PDF)")
        except Exception as e:
            print(f"[akfta] roo_rules upsert failed for {hs_dot}: {e}")

    supabase.table("pipeline_logs").insert({
        "event_type":  "akfta_psr_scraped",
        "source_url":  pdf_url or AKFTA_PAGE,
        "description": f"AKFTA PSR PDF parsed: {len(psr_found)} rules extracted for our HS codes",
        "detail":      {"pdf_url": pdf_url, "rules": psr_found},
    }).execute()
    print(f"[akfta] PSR cross-check complete: {len(psr_found)}/{len(TARGET_HS)} HS codes found in PDF")


# ════════════════════════════════════════════════════════════════
# MITI FTA pages — sync Playwright wrapped in run_in_executor
# ════════════════════════════════════════════════════════════════

async def scrape_miti_fta_rates() -> dict:
    loop = asyncio.get_event_loop()

    # Step 1: update fta_coverage metadata (Playwright, sync in thread)
    fta_result  = await loop.run_in_executor(None, _scrape_miti_coverage_sync)

    # Step 2: download + parse PDFs (httpx, stays async)
    pdf_result  = await scrape_miti_fta_pdfs()

    # Step 3: JKDM sample rates (Playwright, sync in thread)
    jkdm_result = await scrape_jkdm_rates()

    return {
        "ftas_updated":       fta_result.get("ftas_updated", 0),
        "pdf_rates_upserted": pdf_result.get("pdfs_rates_upserted", 0),
        "jkdm_rates_updated": jkdm_result["rates_updated"],
    }


def _scrape_miti_coverage_sync() -> dict:
    """Visit each MITI FTA page and upsert fta_coverage metadata."""
    from playwright.sync_api import sync_playwright
    supabase    = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
    ftas_updated = 0

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page    = browser.new_page()

            for fta_name, url in FTA_PAGES.items():
                meta    = FTA_COVERAGE_META.get(fta_name, {})
                members = FTA_MEMBERS.get(fta_name, [])

                for attempt in range(3):
                    try:
                        page.goto(url, timeout=30000)
                        page.wait_for_load_state("networkidle")

                        supabase.table("fta_coverage").upsert(
                            {
                                "fta_name":         fta_name,
                                "member_countries": members,
                                "source_url":       url,
                                "effective_date":   meta.get("effective_date"),
                                "in_force_date":    meta.get("in_force_date"),
                                "description":      meta.get("description"),
                                "roo_summary":      meta.get("roo_summary"),
                                "last_scraped_at":  datetime.now(timezone.utc).isoformat(),
                            },
                            on_conflict="fta_name"
                        ).execute()

                        print(f"[fta_scraper] Coverage updated: {fta_name}")
                        ftas_updated += 1
                        break

                    except Exception as e:
                        print(f"[fta_scraper] {fta_name} attempt {attempt+1}: {e}")
                        time.sleep(2 ** attempt)
                        if attempt == 2:
                            print(f"[fta_scraper] Skipping {fta_name}")

            browser.close()

    except Exception as e:
        print(f"[fta_scraper] _scrape_miti_coverage_sync failed: {e}")

    return {"ftas_updated": ftas_updated}


def _get_remaining_pdf_links_sync(remaining: dict) -> dict:
    """Use sync Playwright to collect PDF links from MITI pages."""
    from playwright.sync_api import sync_playwright
    pdf_map = {}  # fta_name → [pdf_url, ...]

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--ignore-certificate-errors"])
            page    = browser.new_page()

            for fta_name, page_url in remaining.items():
                try:
                    page.goto(page_url, timeout=30000)
                    page.wait_for_load_state("networkidle")
                    hrefs = page.evaluate(
                        "() => Array.from(document.querySelectorAll('a[href]'))"
                        ".map(a => a.href)"
                        ".filter(h => h.toLowerCase().endsWith('.pdf'))"
                    )
                    if hrefs:
                        pdf_map[fta_name] = hrefs
                        print(f"[fta_pdfs] {fta_name}: found {len(hrefs)} PDF links")
                except Exception as e:
                    print(f"[fta_pdfs] Could not scrape {fta_name} page: {e}")

            browser.close()

    except Exception as e:
        print(f"[fta_pdfs] _get_remaining_pdf_links_sync failed: {e}")

    return pdf_map


# ════════════════════════════════════════════════════════════════
# MITI PDF scraper — httpx (async) + Playwright fallback (sync thread)
# ════════════════════════════════════════════════════════════════

async def scrape_miti_fta_pdfs() -> dict:
    supabase      = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
    total_updated = 0
    total_skipped = 0
    total_errors  = 0

    # Try known directory-listing URLs (no browser needed)
    for fta_name, dir_url in MITI_PDF_DIRS.items():
        pdf_links = await _get_pdf_links_from_directory(dir_url)
        for pdf_url in pdf_links:
            origin = _infer_origin_country(pdf_url, fta_name)
            ok, n  = await _process_pdf(pdf_url, fta_name, origin, supabase)
            if ok:
                total_updated += n
            elif n == -1:
                total_skipped += 1
            else:
                total_errors += 1

    # Fallback: use Playwright (in thread) for pages not in directory list
    scraped_ftas = set(MITI_PDF_DIRS.keys())
    remaining    = {k: v for k, v in FTA_PAGES.items() if k not in scraped_ftas}

    if remaining:
        loop    = asyncio.get_event_loop()
        pdf_map = await loop.run_in_executor(
            None, _get_remaining_pdf_links_sync, remaining
        )
        for fta_name, pdf_links in pdf_map.items():
            for pdf_url in pdf_links:
                origin = _infer_origin_country(pdf_url, fta_name)
                ok, n  = await _process_pdf(pdf_url, fta_name, origin, supabase)
                if ok:
                    total_updated += n
                elif n == -1:
                    total_skipped += 1
                else:
                    total_errors += 1

    print(
        f"[fta_pdfs] Done — {total_updated} rates upserted, "
        f"{total_skipped} PDFs skipped, {total_errors} errors"
    )
    return {
        "pdfs_rates_upserted": total_updated,
        "pdfs_skipped":        total_skipped,
        "pdfs_errors":         total_errors,
    }


async def _get_pdf_links_from_directory(dir_url: str) -> list[str]:
    try:
        async with httpx.AsyncClient(timeout=20, verify=False) as client:
            resp = await client.get(dir_url)
            if resp.status_code != 200:
                return []
            from bs4 import BeautifulSoup
            soup  = BeautifulSoup(resp.text, "html.parser")
            links = []
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if href.lower().endswith(".pdf"):
                    full = href if href.startswith("http") else dir_url.rstrip("/") + "/" + href.lstrip("/")
                    links.append(full)
            return links
    except Exception as e:
        print(f"[fta_pdfs] Directory listing failed {dir_url}: {e}")
        return []


async def _process_pdf(
    pdf_url: str, fta_name: str, origin: str, supabase
) -> tuple[bool, int]:
    try:
        changed, new_hash = await has_source_changed(pdf_url, supabase)
        if not changed:
            return False, -1

        async with httpx.AsyncClient(timeout=60, verify=False) as client:
            resp = await client.get(pdf_url)
            if resp.status_code != 200:
                return False, 0

        records = extract_fta_rates_from_pdf(resp.content, fta_name, origin, pdf_url)
        if not records:
            store_source_hash(pdf_url, new_hash, supabase)
            return True, 0

        count = 0
        for i in range(0, len(records), 100):
            supabase.table("fta_rates").upsert(
                records[i:i + 100], on_conflict="fta_name,hs_code,origin_country"
            ).execute()
            count += len(records[i:i + 100])

        store_source_hash(pdf_url, new_hash, supabase)
        print(f"[fta_pdfs] {fta_name}/{origin}: {count} rates upserted")
        return True, count

    except Exception as e:
        print(f"[fta_pdfs] Error processing {pdf_url}: {e}")
        return False, 0


_COUNTRY_NAME_MAP = {
    # ISO-3 codes (already handled by COUNTRY_CODE_MAP, included here for completeness)
    "CHN": "China",         "JPN": "Japan",          "KOR": "South Korea",
    "AUS": "Australia",     "NZL": "New Zealand",     "IND": "India",
    "CHL": "Chile",         "PAK": "Pakistan",         "TUR": "Turkey",
    "IDN": "Indonesia",     "THA": "Thailand",         "SGP": "Singapore",
    "PHL": "Philippines",   "VNM": "Vietnam",          "BRN": "Brunei",
    "KHM": "Cambodia",      "LAO": "Laos",             "MMR": "Myanmar",
    "HKG": "Hong Kong",     "CAN": "Canada",           "MEX": "Mexico",
    "GBR": "United Kingdom","PER": "Peru",             "BGD": "Bangladesh",
    # Full / variant names found in MITI PDF filenames (uppercase-matched)
    "CHINA":         "China",
    "JAPAN":         "Japan",
    "KOREA":         "South Korea",     "SOUTH-KOREA":  "South Korea",
    "SOUTHKOREA":    "South Korea",     "REPUBLIC-OF-KOREA": "South Korea",
    "AUSTRALIA":     "Australia",
    "NEW-ZEALAND":   "New Zealand",     "NEWZEALAND":   "New Zealand",
    "INDIA":         "India",
    "CHILE":         "Chile",
    "PAKISTAN":      "Pakistan",
    "TURKEY":        "Turkey",
    "INDONESIA":     "Indonesia",
    "THAILAND":      "Thailand",
    "SINGAPORE":     "Singapore",
    "PHILIPPINES":   "Philippines",
    "VIETNAM":       "Vietnam",         "VIET-NAM":     "Vietnam",
    "VIET_NAM":      "Vietnam",
    "BRUNEI":        "Brunei",          "BRUNEI-DARUSSALAM": "Brunei",
    "CAMBODIA":      "Cambodia",
    "LAOS":          "Laos",            "LAO":          "Laos",
    "LAO-PDR":       "Laos",            "LAO_PDR":      "Laos",
    "MYANMAR":       "Myanmar",         "BURMA":        "Myanmar",
    "HONG-KONG":     "Hong Kong",       "HONGKONG":     "Hong Kong",
    "CANADA":        "Canada",
    "MEXICO":        "Mexico",
    "UNITED-KINGDOM":"United Kingdom",  "UK":           "United Kingdom",
    "PERU":          "Peru",
    "BANGLADESH":    "Bangladesh",
    "SAUDI-ARABIA":  "Saudi Arabia",    "SAUDIARABIA":  "Saudi Arabia",
    "IRAN":          "Iran",
    "EGYPT":         "Egypt",
    "MOROCCO":       "Morocco",
    "NIGERIA":       "Nigeria",
    "JORDAN":        "Jordan",
    "TUNISIA":       "Tunisia",
    "SENEGAL":       "Senegal",
}


def _infer_origin_country(url: str, fta_name: str) -> str:
    """
    Infer origin country from PDF filename.
    Tries ISO-3 codes and full/variant country names (case-insensitive).
    Falls back to bilateral default for bilateral FTAs, else 'Unknown'.
    """
    filename = url.split("/")[-1].upper().replace(".PDF", "").replace("_", "-")

    # Try longest match first to avoid "KOR" matching before "SOUTH-KOREA"
    for token in sorted(_COUNTRY_NAME_MAP.keys(), key=len, reverse=True):
        if token in filename:
            return _COUNTRY_NAME_MAP[token]

    bilateral_defaults = {
        "MAFTA": "Australia", "MCFTA": "Chile",       "MICECA": "India",
        "MJEPA": "Japan",     "MNZFTA": "New Zealand", "MPCEPA": "Pakistan",
        "MTFTA": "Turkey",
    }
    country = bilateral_defaults.get(fta_name)
    if country:
        return country

    print(f"[fta_scraper] Could not infer origin country from: {url} (fta={fta_name})")
    return "Unknown"


# ════════════════════════════════════════════════════════════════
# Seed (offline fallback + one-time metadata fix)
# ════════════════════════════════════════════════════════════════

async def seed_fta_data() -> None:
    try:
        supabase = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)

        for fta_name, members in FTA_MEMBERS.items():
            meta = FTA_COVERAGE_META.get(fta_name, {})
            supabase.table("fta_coverage").upsert(
                {
                    "fta_name":         fta_name,
                    "member_countries": members,
                    "source_url":       meta.get("source_url"),
                    "effective_date":   meta.get("effective_date"),
                    "in_force_date":    meta.get("in_force_date"),
                    "description":      meta.get("description"),
                    "roo_summary":      meta.get("roo_summary"),
                },
                on_conflict="fta_name"
            ).execute()

        fta_rates = [
            {
                "fta_name": "RCEP", "hs_code": "8534.00", "origin_country": "China",
                "preferential_rate_pct": 0.0, "effective_date": "2022-01-01",
                "rate_staging": {"2022": 0.0}, "final_rate": 0.0, "final_year": 2022,
                "staging_category": "EIF",
            },
            {
                "fta_name": "RCEP", "hs_code": "8542.31", "origin_country": "China",
                "preferential_rate_pct": 0.0, "effective_date": "2022-01-01",
                "rate_staging": {"2022": 0.0}, "final_rate": 0.0, "final_year": 2022,
                "staging_category": "EIF",
            },
            {
                "fta_name": "RCEP", "hs_code": "8501.52", "origin_country": "Japan",
                "preferential_rate_pct": 0.0, "effective_date": "2022-01-01",
                "rate_staging": {"2022": 0.0}, "final_rate": 0.0, "final_year": 2022,
                "staging_category": "EIF",
            },
            {
                "fta_name": "RCEP", "hs_code": "7408.11", "origin_country": "China",
                "preferential_rate_pct": 1.0, "effective_date": "2022-01-01",
                "rate_staging": {
                    "2022": 5.0, "2023": 4.0, "2024": 3.0,
                    "2025": 2.0, "2026": 1.0, "2027": 0.0,
                },
                "final_rate": 0.0, "final_year": 2027, "staging_category": "SL-5",
            },
            {
                "fta_name": "CPTPP", "hs_code": "8534.00", "origin_country": "Japan",
                "preferential_rate_pct": 0.0, "effective_date": "2022-11-29",
                "rate_staging": {"2022": 0.0}, "final_rate": 0.0, "final_year": 2022,
                "staging_category": "EIF",
            },
            {
                "fta_name": "CPTPP", "hs_code": "7604.29", "origin_country": "Australia",
                "preferential_rate_pct": 1.0, "effective_date": "2022-11-29",
                "rate_staging": {
                    "2023": 4.0, "2024": 3.0, "2025": 2.0,
                    "2026": 1.0, "2027": 0.0,
                },
                "final_rate": 0.0, "final_year": 2027, "staging_category": "B5",
            },
            {
                "fta_name": "ACFTA",  "hs_code": "8534.00", "origin_country": "China",
                "preferential_rate_pct": 0.0, "effective_date": "2005-07-20",
                "rate_staging": None, "final_rate": 0.0,
            },
            {
                "fta_name": "AKFTA",  "hs_code": "8501.52", "origin_country": "South Korea",
                "preferential_rate_pct": 0.0, "effective_date": "2007-06-01",
                "rate_staging": None, "final_rate": 0.0,
            },
            {
                "fta_name": "ATIGA",  "hs_code": "7408.11", "origin_country": "Vietnam",
                "preferential_rate_pct": 0.0, "effective_date": "2010-01-01",
                "rate_staging": None, "final_rate": 0.0,
            },
            {
                "fta_name": "AKFTA",  "hs_code": "9001.90", "origin_country": "South Korea",
                "preferential_rate_pct": 0.0, "effective_date": "2007-06-01",
                "rate_staging": None, "final_rate": 0.0,
            },
            {
                "fta_name": "AJCEP",  "hs_code": "8525.80", "origin_country": "Japan",
                "preferential_rate_pct": 0.0, "effective_date": "2008-12-01",
                "rate_staging": None, "final_rate": 0.0,
            },
            {
                "fta_name": "PDK", "hs_code": "8534.00", "origin_country": "MFN",
                "preferential_rate_pct": 5.0, "effective_date": "2025-01-01",
                "rate_staging": None, "final_rate": 5.0,
            },
            {
                "fta_name": "PDK", "hs_code": "8501.52", "origin_country": "MFN",
                "preferential_rate_pct": 5.0, "effective_date": "2025-01-01",
                "rate_staging": None, "final_rate": 5.0,
            },
            {
                "fta_name": "PDK", "hs_code": "7408.11", "origin_country": "MFN",
                "preferential_rate_pct": 5.0, "effective_date": "2025-01-01",
                "rate_staging": None, "final_rate": 5.0,
            },
            {
                "fta_name": "PDK", "hs_code": "7604.29", "origin_country": "MFN",
                "preferential_rate_pct": 5.0, "effective_date": "2025-01-01",
                "rate_staging": None, "final_rate": 5.0,
            },
            {
                "fta_name": "PDK", "hs_code": "9001.90", "origin_country": "MFN",
                "preferential_rate_pct": 0.0, "effective_date": "2025-01-01",
                "rate_staging": None, "final_rate": 0.0,
            },
        ]

        for rate in fta_rates:
            supabase.table("fta_rates").upsert(
                rate, on_conflict="fta_name,hs_code,origin_country"
            ).execute()

        roo_rules = [
            {"fta_name": "ACFTA",  "hs_code": "8534.00", "roo_type": "RVC", "rvc_threshold_pct": 40.0},
            {"fta_name": "AKFTA",  "hs_code": "8501.52", "roo_type": "RVC", "rvc_threshold_pct": 40.0},
            {"fta_name": "ATIGA",  "hs_code": "7408.11", "roo_type": "RVC", "rvc_threshold_pct": 40.0},
            {"fta_name": "RCEP",   "hs_code": "7604.29", "roo_type": "tariff_shift",
             "tariff_shift_description": "CTSH - Change in Tariff Sub-Heading"},
            {"fta_name": "AKFTA",  "hs_code": "9001.90", "roo_type": "RVC", "rvc_threshold_pct": 40.0},
            {"fta_name": "RCEP",   "hs_code": "8534.00", "roo_type": "RVC", "rvc_threshold_pct": 40.0},
            {"fta_name": "CPTPP",  "hs_code": "8534.00", "roo_type": "RVC", "rvc_threshold_pct": 40.0},
        ]

        for rule in roo_rules:
            supabase.table("roo_rules").upsert(
                rule, on_conflict="fta_name,hs_code"
            ).execute()

        print("[fta_scraper] seed_fta_data completed — all 17 FTAs with metadata")

    except Exception as e:
        print(f"[fta_scraper] seed_fta_data failed: {e}")
